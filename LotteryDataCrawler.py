import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import time
import random
from datetime import datetime
import os
from openpyxl import load_workbook
from tqdm import tqdm
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.filters import FilterColumn, Filters


class LotteryDataCrawler:
    def __init__(self):
        # 更完整的请求头，模拟 Chrome 浏览器
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Cache-Control': 'max-age=0',
        }
        self.session = requests.Session()
        self.session.headers.update(self.headers)
        
        # 添加重试策略
        adapter = requests.adapters.HTTPAdapter(
            max_retries=3,
            pool_connections=10,
            pool_maxsize=10
        )
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)
        
        self._warmed_up = False  # 标记是否已预热（获取cookies）

    def _warmup(self):
        """预热 session，获取 78500.cn 的 cookies"""
        if self._warmed_up:
            return
        try:
            # 先访问首页获取 cookies
            self.session.get("https://kaijiang.78500.cn/", timeout=10)
            time.sleep(random.uniform(1, 2))
            self._warmed_up = True
        except Exception as e:
            print(f"预热失败: {e}")

    # 大乐透函数（保持不变）
    def get_dlt_history_data_all(self):
        """获取所有大乐透历史开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = "https://datachart.500.com/dlt/history/history.shtml"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 15:
                    continue
                period = cols[0].text.strip()
                front_nums = [cols[i].text.strip().zfill(2) for i in range(1, 6)]
                back_nums = [cols[i].text.strip().zfill(2) for i in range(6, 8)]
                prize_pool = cols[8].text.strip()
                first_prize_count = cols[9].text.strip()
                first_prize_amount = cols[10].text.strip()
                second_prize_count = cols[11].text.strip()
                second_prize_amount = cols[12].text.strip()
                total_bet = cols[13].text.strip()
                date = cols[14].text.strip()
                data.append([
                    period,
                    front_nums[0], front_nums[1], front_nums[2], front_nums[3], front_nums[4],
                    back_nums[0], back_nums[1],
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    def get_dlt_history_data_by_range(self, start_period, end_period):
        """按期号范围获取大乐透历史开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = f"https://datachart.500.com/dlt/history/newinc/history.php?start={start_period}&end={end_period}"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 15:
                    continue
                period = cols[0].text.strip()
                front_nums = [cols[i].text.strip().zfill(2) for i in range(1, 6)]
                back_nums = [cols[i].text.strip().zfill(2) for i in range(6, 8)]
                prize_pool = cols[8].text.strip()
                first_prize_count = cols[9].text.strip()
                first_prize_amount = cols[10].text.strip()
                second_prize_count = cols[11].text.strip()
                second_prize_amount = cols[12].text.strip()
                total_bet = cols[13].text.strip()
                date = cols[14].text.strip()
                data.append([
                    period,
                    front_nums[0], front_nums[1], front_nums[2], front_nums[3], front_nums[4],
                    back_nums[0], back_nums[1],
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    def get_all_dlt_history(self):
        """获取所有大乐透历史数据，分批获取然后合并"""
        all_data = self.get_dlt_history_data_all()
        if all_data is not None and len(all_data) > 2700:
            return all_data
        all_data = pd.DataFrame()
        batches = [
            (1, 1000), (1001, 2000), (2001, 3000), (3001, 4000), (4001, 5000),
            (5001, 6000), (6001, 7000), (7001, 8000), (8001, 9000), (9001, 10000),
            (10001, 11000), (11001, 12000), (12001, 13000), (13001, 14000), (14001, 15000),
            (15001, 16000), (16001, 17000), (17001, 18000), (18001, 19000), (19001, 20000),
            (20001, 21000), (21001, 22000), (22001, 23000), (23001, 24000), (24001, 25000),
            (25001, 26000), (26001, 27000), (27001, 28000), (28001, 29000), (29001, 30000)
        ]
        for start, end in tqdm(batches, desc="获取大乐透历史数据"):
            batch_data = self.get_dlt_history_data_by_range(start, end)
            if batch_data is not None and len(batch_data) > 0:
                all_data = pd.concat([all_data, batch_data], ignore_index=True)
            time.sleep(random.uniform(1, 3))
        if not all_data.empty:
            all_data['开奖日期'] = pd.to_datetime(all_data['开奖日期'])
            all_data = all_data.sort_values(by='开奖日期', ascending=True)
            all_data = all_data.reset_index(drop=True)
            for col in ['期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']:
                all_data[col] = all_data[col].astype(str)
            all_data['开奖日期'] = all_data['开奖日期'].dt.strftime('%Y-%m-%d')
        return all_data

    def get_latest_dlt_data(self):
        """获取最新的大乐透开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = "https://datachart.500.com/dlt/history/newinc/history.php?start=00001&end=99999"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 15:
                    continue
                period = cols[0].text.strip()
                front_nums = [cols[i].text.strip().zfill(2) for i in range(1, 6)]
                back_nums = [cols[i].text.strip().zfill(2) for i in range(6, 8)]
                prize_pool = cols[8].text.strip()
                first_prize_count = cols[9].text.strip()
                first_prize_amount = cols[10].text.strip()
                second_prize_count = cols[11].text.strip()
                second_prize_amount = cols[12].text.strip()
                total_bet = cols[13].text.strip()
                date = cols[14].text.strip()
                data.append([
                    period,
                    front_nums[0], front_nums[1], front_nums[2], front_nums[3], front_nums[4],
                    back_nums[0], back_nums[1],
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    # 双色球函数（保持不变）
    def get_ssq_history_data_all(self):
        """获取所有双色球历史开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = "https://datachart.500.com/ssq/history/history.shtml"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 16:
                    continue
                period = cols[0].text.strip()
                red_nums = [cols[i].text.strip().zfill(2) for i in range(1, 7)]
                blue_num = cols[7].text.strip().zfill(2)
                prize_pool = cols[9].text.strip()
                first_prize_count = cols[10].text.strip()
                first_prize_amount = cols[11].text.strip()
                second_prize_count = cols[12].text.strip()
                second_prize_amount = cols[13].text.strip()
                total_bet = cols[14].text.strip()
                date = cols[15].text.strip()
                data.append([
                    period,
                    red_nums[0], red_nums[1], red_nums[2], red_nums[3], red_nums[4], red_nums[5],
                    blue_num,
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    def get_ssq_history_data_by_range(self, start_period, end_period):
        """按期号范围获取双色球历史开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = f"https://datachart.500.com/ssq/history/newinc/history.php?start={start_period}&end={end_period}"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 16:
                    continue
                period = cols[0].text.strip()
                red_nums = [cols[i].text.strip().zfill(2) for i in range(1, 7)]
                blue_num = cols[7].text.strip().zfill(2)
                prize_pool = cols[9].text.strip()
                first_prize_count = cols[10].text.strip()
                first_prize_amount = cols[11].text.strip()
                second_prize_count = cols[12].text.strip()
                second_prize_amount = cols[13].text.strip()
                total_bet = cols[14].text.strip()
                date = cols[15].text.strip()
                data.append([
                    period,
                    red_nums[0], red_nums[1], red_nums[2], red_nums[3], red_nums[4], red_nums[5],
                    blue_num,
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    def get_all_ssq_history(self):
        """获取所有双色球历史数据，分批获取然后合并"""
        all_data = self.get_ssq_history_data_all()
        if all_data is not None and len(all_data) > 2000:
            return all_data
        all_data = pd.DataFrame()
        batches = [
            (1, 1000), (1001, 2000), (2001, 3000), (3001, 4000), (4001, 5000),
            (5001, 6000), (6001, 7000), (7001, 8000), (8001, 9000), (9001, 10000),
            (10001, 11000), (11001, 12000), (12001, 13000), (13001, 14000), (14001, 15000),
            (15001, 16000), (16001, 17000), (17001, 18000), (18001, 19000), (19001, 20000),
            (20001, 21000), (21001, 22000), (22001, 23000), (23001, 24000), (24001, 25000),
            (25001, 26000)
        ]
        for start, end in tqdm(batches, desc="获取双色球历史数据"):
            batch_data = self.get_ssq_history_data_by_range(start, end)
            if batch_data is not None and len(batch_data) > 0:
                all_data = pd.concat([all_data, batch_data], ignore_index=True)
            time.sleep(random.uniform(1, 3))
        if not all_data.empty:
            all_data['开奖日期'] = pd.to_datetime(all_data['开奖日期'])
            all_data = all_data.sort_values(by='开奖日期', ascending=True)
            all_data = all_data.reset_index(drop=True)
            for col in ['期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球']:
                all_data[col] = all_data[col].astype(str)
            all_data['开奖日期'] = all_data['开奖日期'].dt.strftime('%Y-%m-%d')
        return all_data

    def get_latest_ssq_data(self):
        """获取最新的双色球开奖数据"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        }
        url = "https://datachart.500.com/ssq/history/newinc/history.php?start=00001&end=99999"
        try:
            response = requests.get(url, headers=headers, timeout=30)
            response.encoding = 'gb2312'
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'id': 'tablelist'})
            if not table:
                return None
            data = []
            rows = table.find_all('tr')[2:]
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 16:
                    continue
                period = cols[0].text.strip()
                red_nums = [cols[i].text.strip().zfill(2) for i in range(1, 7)]
                blue_num = cols[7].text.strip().zfill(2)
                prize_pool = cols[9].text.strip()
                first_prize_count = cols[10].text.strip()
                first_prize_amount = cols[11].text.strip()
                second_prize_count = cols[12].text.strip()
                second_prize_amount = cols[13].text.strip()
                total_bet = cols[14].text.strip()
                date = cols[15].text.strip()
                data.append([
                    period,
                    red_nums[0], red_nums[1], red_nums[2], red_nums[3], red_nums[4], red_nums[5],
                    blue_num,
                    prize_pool,
                    first_prize_count,
                    first_prize_amount,
                    second_prize_count,
                    second_prize_amount,
                    total_bet,
                    date
                ])
            columns = [
                '期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球',
                '奖池奖金(元)', '一等奖注数', '一等奖奖金(元)', '二等奖注数', '二等奖奖金(元)',
                '总投注额(元)', '开奖日期'
            ]
            df = pd.DataFrame(data, columns=columns)
            df['开奖日期'] = pd.to_datetime(df['开奖日期'])
            df = df.sort_values(by='开奖日期', ascending=True)
            df = df.reset_index(drop=True)
            for col in ['期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球']:
                df[col] = df[col].astype(str)
            df['开奖日期'] = df['开奖日期'].dt.strftime('%Y-%m-%d')
            return df
        except Exception:
            return None

    # 排列五函数（优化版）
    def get_pl5_recent_data(self, count=30):
        """获取排列五最近N期数据（带预热和重试）"""
        self._warmup()  # 确保已获取cookies

        base_url = "https://kaijiang.78500.cn/p5/"
        current_year = datetime.now().year
        params = {
            'action': 'years',
            'year': current_year
        }
        
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                response = self.session.get(base_url, params=params, timeout=15)
                response.encoding = 'gb2312'
                if response.status_code == 200:
                    df = self.parse_pl5_html(response.text)
                    if df is not None and len(df) > 0:
                        # 取最近count期数据
                        recent_df = df.head(count).copy()
                        # 按时间顺序排列（从早到晚）
                        recent_df = recent_df.sort_values('开奖日期', ascending=True).reset_index(drop=True)
                        return recent_df
                    else:
                        # 解析失败可能是IP被屏蔽，等待后重试
                        time.sleep(random.uniform(3, 6))
                else:
                    time.sleep(random.uniform(2, 4))
            except Exception as e:
                print(f"排列五请求异常 (尝试 {attempt+1}/{max_attempts}): {e}")
                time.sleep(random.uniform(5, 10))
        return None

    def parse_pl5_html(self, html_content):
        """
        解析排列五HTML内容，提取开奖数据
        根据实际页面结构修正列索引：
        - 期号: cols[0]
        - 开奖日期: cols[1]
        - 开奖号码: cols[2] (内含<a>标签，文本带空格)
        - 直选注数(一等奖注数): cols[3]
        - 直选奖金: cols[4] (暂不使用)
        - 销售额: cols[5]
        - 奖池奖金: cols[6] (暂不使用)
        - 详情链接: cols[7] (忽略)
        """
        soup = BeautifulSoup(html_content, 'html.parser')
        table = soup.find('table', class_='kjls')
        if not table:
            return None

        data = []
        # 跳过表头行（前两行）
        rows = table.find_all('tr')[2:]

        for row in rows:
            cols = row.find_all('td')
            # 实际数据行应有8个td
            if len(cols) < 8:
                continue
            try:
                # 期号
                period = cols[0].get_text(strip=True)
                if not period.isdigit():
                    continue

                # 开奖日期
                date_text = cols[1].get_text(strip=True)

                # 开奖号码：优先取<a>标签内的文本，否则直接取td文本
                number_cell = cols[2]
                number_link = number_cell.find('a')
                if number_link:
                    numbers = number_link.get_text(strip=True)
                else:
                    numbers = number_cell.get_text(strip=True)
                # 保留原始格式（带空格）
                numbers = numbers.strip()
                # 验证号码格式（应为5位数字，可能有空格）
                clean_numbers = numbers.replace(' ', '')
                if not re.match(r'^\d{5}$', clean_numbers):
                    continue

                # 直选注数（一等奖注数）
                first_prize_count = self.parse_prize_count(cols[3].get_text(strip=True))

                # 销售额
                sales_text = cols[5].get_text(strip=True)
                sales_amount = self.parse_sales_amount(sales_text)

                data.append({
                    '期号': period,
                    '开奖日期': date_text,
                    '销售金额': sales_amount,
                    '开奖号码': numbers,
                    '一等奖注数': first_prize_count
                })
            except Exception:
                continue

        return pd.DataFrame(data) if data else None

    # 福彩3D函数（优化版）
    def get_3d_recent_data(self, count=30):
        """获取福彩3D最近N期数据（带预热和重试）"""
        self._warmup()

        base_url = "https://kaijiang.78500.cn/3d/"
        current_year = datetime.now().year
        params = {
            'action': 'years',
            'year': current_year
        }
        
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                response = self.session.get(base_url, params=params, timeout=15)
                response.encoding = 'gb2312'
                if response.status_code == 200:
                    df = self.parse_3d_html(response.text)
                    if df is not None and len(df) > 0:
                        recent_df = df.head(count).copy()
                        recent_df = recent_df.sort_values('开奖日期', ascending=True).reset_index(drop=True)
                        return recent_df
                    else:
                        time.sleep(random.uniform(3, 6))
                else:
                    time.sleep(random.uniform(2, 4))
            except Exception as e:
                print(f"福彩3D请求异常 (尝试 {attempt+1}/{max_attempts}): {e}")
                time.sleep(random.uniform(5, 10))
        return None

    def parse_3d_html(self, html_content):
        """
        解析福彩3D HTML内容，提取开奖数据
        列索引参考实际页面：
        - 期号: cols[0]
        - 开奖日期: cols[1]
        - 销售额: cols[2]
        - 开奖号码: cols[3] (内含<a>标签)
        """
        soup = BeautifulSoup(html_content, 'html.parser')
        table = soup.find('table', class_='kjls')
        if not table:
            return None

        data = []
        rows = table.find_all('tr')[2:]

        for row in rows:
            cols = row.find_all('td')
            if len(cols) >= 9:  # 福彩3D有更多列，但我们只需要前几个
                try:
                    period = cols[0].get_text(strip=True)
                    if not period.isdigit():
                        continue

                    date_text = cols[1].get_text(strip=True)

                    sales_text = cols[2].get_text(strip=True)
                    sales_amount = self.parse_sales_amount(sales_text)

                    number_cell = cols[3]
                    number_link = number_cell.find('a')
                    if number_link:
                        numbers = number_link.get_text(strip=True)
                    else:
                        numbers = number_cell.get_text(strip=True)

                    numbers = numbers.strip()
                    clean_numbers = numbers.replace(' ', '')
                    if not re.match(r'^\d{3}$', clean_numbers):
                        continue

                    data.append({
                        '期号': period,
                        '开奖日期': date_text,
                        '销售金额': sales_amount,
                        '开奖号码': numbers
                    })
                except Exception:
                    continue

        return pd.DataFrame(data) if data else None

    def parse_sales_amount(self, text):
        """解析销售金额"""
        if text in ['', '-', '0']:
            return 0
        try:
            return int(text.replace(',', ''))
        except:
            return 0

    def parse_prize_count(self, text):
        """解析奖注数"""
        if text in ['', '-']:
            return 0
        try:
            return int(text)
        except:
            return 0


# 通用函数（保持不变）
def get_existing_data(filename='Tools.xlsx', sheet_name='dltall'):
    """从Excel文件中读取现有的开奖数据"""
    if not os.path.exists(filename):
        return pd.DataFrame()
    try:
        existing_df = pd.read_excel(filename, sheet_name=sheet_name, dtype=str)
        if '开奖日期' in existing_df.columns:
            existing_df['开奖日期'] = pd.to_datetime(existing_df['开奖日期'], errors='coerce').dt.strftime('%Y-%m-%d')
        if sheet_name == 'dltall':
            for col in ['期号', '前区1', '前区2', '前区3', '前区4', '前区5', '后区1', '后区2']:
                if col in existing_df.columns:
                    existing_df[col] = existing_df[col].astype(str).apply(lambda x: x.zfill(2) if x.isdigit() and len(x) == 1 else x)
        elif sheet_name == 'ssq':
            for col in ['期号', '红球1', '红球2', '红球3', '红球4', '红球5', '红球6', '蓝球']:
                if col in existing_df.columns:
                    existing_df[col] = existing_df[col].astype(str).apply(lambda x: x.zfill(2) if x.isdigit() and len(x) == 1 else x)
        return existing_df
    except Exception:
        return pd.DataFrame()

def safe_save_to_excel(df, filename='Tools.xlsx', sheet_name='dltall'):
    """安全保存数据，不修改现有文件结构，保留外部链接和其他工作表内容"""
    if df is None or len(df) == 0:
        return
    try:
        temp_filename = f"temp_{filename}"
        with pd.ExcelWriter(temp_filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        original_book = load_workbook(filename)
        temp_book = load_workbook(temp_filename)
        temp_sheet = temp_book[sheet_name]
        new_data = []
        for row in temp_sheet.iter_rows(values_only=True):
            new_data.append(row)
        if sheet_name in original_book.sheetnames:
            original_sheet = original_book[sheet_name]
            for row_idx in range(2, original_sheet.max_row + 1):
                for col_idx in range(1, original_sheet.max_column + 1):
                    original_sheet.cell(row=row_idx, column=col_idx).value = None
            for row_idx, row_data in enumerate(new_data[1:], 2):
                for col_idx, value in enumerate(row_data, 1):
                    if col_idx <= original_sheet.max_column:
                        original_sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            original_sheet = original_book.create_sheet(sheet_name)
            for row_idx, row_data in enumerate(new_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    original_sheet.cell(row=row_idx, column=col_idx, value=value)
        original_book.save(filename)
        os.remove(temp_filename)
    except Exception:
        try:
            if os.path.exists(filename):
                book = load_workbook(filename)
                if sheet_name in book.sheetnames:
                    std = book[sheet_name]
                    book.remove(std)
                sheet = book.create_sheet(sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)
                book.save(filename)
            else:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception:
            pass

def get_existing_dates(filename, sheet_name):
    """获取工作表中已存在的开奖日期"""
    try:
        if not os.path.exists(filename):
            return set()
        existing_df = pd.read_excel(filename, sheet_name=sheet_name)
        if existing_df.empty or '开奖日期' not in existing_df.columns:
            return set()
        return set(existing_df['开奖日期'].astype(str))
    except Exception:
        return set()

def incremental_save_to_excel(new_df, filename='Tools.xlsx', sheet_name='pl5'):
    """增量保存数据到Excel，基于开奖日期去重"""
    try:
        existing_dates = get_existing_dates(filename, sheet_name)
        new_rows = []
        for _, row in new_df.iterrows():
            if str(row['开奖日期']) not in existing_dates:
                new_rows.append(row)
        if not new_rows:
            return True
        new_data_df = pd.DataFrame(new_rows)
        if os.path.exists(filename):
            workbook = load_workbook(filename)
            if sheet_name in workbook.sheetnames:
                existing_df = pd.read_excel(filename, sheet_name=sheet_name)
                combined_df = pd.concat([existing_df, new_data_df], ignore_index=True)
                combined_df = combined_df.sort_values('开奖日期', ascending=True).reset_index(drop=True)
            else:
                combined_df = new_data_df
        else:
            workbook = load_workbook()
            combined_df = new_data_df
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)
        worksheet.delete_rows(1, worksheet.max_row)
        headers = list(combined_df.columns)
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)
        for row_idx, row in enumerate(combined_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        worksheet.freeze_panes = 'A2'
        workbook.save(filename)
        workbook.close()
        return True
    except Exception:
        return False

def format_pl5_dataframe(df):
    """格式化排列五数据框"""
    expected_columns = ['期号', '开奖日期', '销售金额', '开奖号码', '一等奖注数']
    df = df[expected_columns].copy()
    return df

def format_3d_dataframe(df):
    """格式化福彩3D数据框"""
    expected_columns = ['期号', '开奖日期', '销售金额', '开奖号码']
    df = df[expected_columns].copy()
    return df

def update_ev_sheet_with_lottery_data(filename='Tools.xlsx'):
    """将pl5和3D工作表的开奖数据自动写入EV工作表的相应列"""
    try:
        if not os.path.exists(filename):
            print("文件不存在")
            return False
        workbook = load_workbook(filename)
        required_sheets = ['pl5', '3D', 'EV']
        for sheet in required_sheets:
            if sheet not in workbook.sheetnames:
                print(f"缺少必要的工作表: {sheet}")
                workbook.close()
                return False
        pl5_df = pd.read_excel(filename, sheet_name='pl5')
        d3_df = pd.read_excel(filename, sheet_name='3D')
        ev_ws = workbook['EV']
        print(f"pl5数据行数: {len(pl5_df)}")
        print(f"3D数据行数: {len(d3_df)}")
        print(f"EV工作表行数: {ev_ws.max_row}")
        date_to_row = {}
        for row in range(2, ev_ws.max_row + 1):
            date_cell = ev_ws.cell(row, 1).value
            if date_cell is None:
                continue
            if isinstance(date_cell, datetime):
                date_str = date_cell.strftime('%Y-%m-%d')
            else:
                date_str = str(date_cell).strip()
                if ' ' in date_str:
                    date_str = date_str.split(' ')[0]
            date_to_row[date_str] = row
        print(f"EV工作表日期映射数量: {len(date_to_row)}")
        pl5_updated = 0
        for _, row in pl5_df.iterrows():
            date_str = str(row['开奖日期']).strip()
            if date_str in date_to_row:
                ev_row = date_to_row[date_str]
                b_col = 2
                cell_value = ev_ws.cell(ev_row, b_col).value
                if cell_value is None or cell_value == '':
                    numbers = str(row['开奖号码']).strip()
                    clean_numbers = numbers.replace(' ', '')
                    if len(clean_numbers) == 5:
                        if ' ' not in numbers:
                            numbers = ' '.join(clean_numbers)
                        ev_ws.cell(ev_row, b_col).value = numbers
                        pl5_updated += 1
                        print(f"更新PL5数据: 日期 {date_str}, 号码 {numbers}")
        d3_updated = 0
        for _, row in d3_df.iterrows():
            date_str = str(row['开奖日期']).strip()
            if date_str in date_to_row:
                ev_row = date_to_row[date_str]
                e_col = 5
                cell_value = ev_ws.cell(ev_row, e_col).value
                if cell_value is None or cell_value == '':
                    numbers = str(row['开奖号码']).strip()
                    clean_numbers = numbers.replace(' ', '')
                    if len(clean_numbers) == 3:
                        if ' ' not in numbers:
                            numbers = ' '.join(clean_numbers)
                        ev_ws.cell(ev_row, e_col).value = numbers
                        d3_updated += 1
                        print(f"更新3D数据: 日期 {date_str}, 号码 {numbers}")
        workbook.save(filename)
        workbook.close()
        print(f"PL5更新了 {pl5_updated} 条数据")
        print(f"3D更新了 {d3_updated} 条数据")
        return True
    except Exception as e:
        print(f"更新EV工作表时出错: {str(e)}")
        if 'workbook' in locals():
            workbook.close()
        return False

def update_dlt_data(crawler):
    """更新大乐透数据：只获取并添加新期数"""
    existing_df = get_existing_data('Tools.xlsx', 'dltall')
    if existing_df.empty:
        with tqdm(total=1, desc="获取大乐透历史数据") as pbar:
            all_data = crawler.get_all_dlt_history()
            pbar.update(1)
        if all_data is not None and len(all_data) > 0:
            safe_save_to_excel(all_data, 'Tools.xlsx', 'dltall')
        return
    latest_data = crawler.get_latest_dlt_data()
    if latest_data is None or latest_data.empty:
        return
    new_data = latest_data[~latest_data['期号'].isin(existing_df['期号'])]
    if new_data.empty:
        return
    if '开奖日期' in existing_df.columns:
        existing_df['开奖日期'] = pd.to_datetime(existing_df['开奖日期'], errors='coerce').dt.strftime('%Y-%m-%d')
    combined_data = pd.concat([existing_df, new_data], ignore_index=True)
    combined_data['开奖日期'] = pd.to_datetime(combined_data['开奖日期'], errors='coerce')
    combined_data = combined_data.dropna(subset=['开奖日期'])
    combined_data = combined_data.sort_values(by='开奖日期', ascending=True)
    combined_data = combined_data.reset_index(drop=True)
    combined_data['开奖日期'] = combined_data['开奖日期'].dt.strftime('%Y-%m-%d')
    safe_save_to_excel(combined_data, 'Tools.xlsx', 'dltall')

def update_ssq_data(crawler):
    """更新双色球数据：只获取并添加新期数"""
    existing_df = get_existing_data('Tools.xlsx', 'ssq')
    if existing_df.empty:
        with tqdm(total=1, desc="获取双色球历史数据") as pbar:
            all_data = crawler.get_all_ssq_history()
            pbar.update(1)
        if all_data is not None and len(all_data) > 0:
            safe_save_to_excel(all_data, 'Tools.xlsx', 'ssq')
        return
    latest_data = crawler.get_latest_ssq_data()
    if latest_data is None or latest_data.empty:
        return
    new_data = latest_data[~latest_data['期号'].isin(existing_df['期号'])]
    if new_data.empty:
        return
    if '开奖日期' in existing_df.columns:
        existing_df['开奖日期'] = pd.to_datetime(existing_df['开奖日期'], errors='coerce').dt.strftime('%Y-%m-%d')
    combined_data = pd.concat([existing_df, new_data], ignore_index=True)
    combined_data['开奖日期'] = pd.to_datetime(combined_data['开奖日期'], errors='coerce')
    combined_data = combined_data.dropna(subset=['开奖日期'])
    combined_data = combined_data.sort_values(by='开奖日期', ascending=True)
    combined_data = combined_data.reset_index(drop=True)
    combined_data['开奖日期'] = combined_data['开奖日期'].dt.strftime('%Y-%m-%d')
    safe_save_to_excel(combined_data, 'Tools.xlsx', 'ssq')

def update_pl5_data(crawler):
    """更新排列五数据"""
    pl5_df = crawler.get_pl5_recent_data(30)
    if pl5_df is not None and len(pl5_df) > 0:
        pl5_formatted = format_pl5_dataframe(pl5_df)
        incremental_save_to_excel(pl5_formatted, 'Tools.xlsx', 'pl5')

def update_3d_data(crawler):
    """更新福彩3D数据"""
    d3_df = crawler.get_3d_recent_data(30)
    if d3_df is not None and len(d3_df) > 0:
        d3_formatted = format_3d_dataframe(d3_df)
        incremental_save_to_excel(d3_formatted, 'Tools.xlsx', '3D')

def main():
    """主函数 - 更新所有彩票数据"""
    print("彩票开奖数据获取工具 - 完整版 (优化版)")
    print("=" * 50)
    crawler = LotteryDataCrawler()
    print("正在更新大乐透数据...")
    update_dlt_data(crawler)
    print("正在更新双色球数据...")
    update_ssq_data(crawler)
    print("正在更新排列五数据...")
    update_pl5_data(crawler)
    print("正在更新福彩3D数据...")
    update_3d_data(crawler)
    print("正在将PL5和3D数据更新到EV工作表...")
    success = update_ev_sheet_with_lottery_data('Tools.xlsx')
    if success:
        print("EV工作表更新成功!")
    else:
        print("EV工作表更新失败!")
    file_path = os.path.abspath('Tools.xlsx')
    print(f"文件保存路径: {file_path}")
    print("所有数据更新完成!")

if __name__ == "__main__":
    main()